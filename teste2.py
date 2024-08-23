import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime

Planilha_canhoto = pd.read_excel("Base_canhotos.xlsx")
#print(Planilha_canhoto)

status = datetime.now()
status = status.strftime("%d/%m")
status = f'Enviado {status}'
#print(status)


numero_linhas = len(Planilha_canhoto)
#print(numero_linhas)
linha_especifica = 1
for i, linha in enumerate(Planilha_canhoto.index):
    cte = Planilha_canhoto.loc[linha, "Ct-e Online"]
    cte = int(cte)
    print(cte)
    protocolo = 21826
    linha_especifica += 1 
    caminho_do_arquivo = 'Base_canhotos.xlsx'
    nome_da_aba = 'Plan1'
    wb = load_workbook(caminho_do_arquivo)
    ws = wb[nome_da_aba]
    coluna_lancamento = 'B'  
    if linha > ws.max_row:
        ws[coluna_lancamento + str(linha_especifica)] = protocolo
    else:
        ws[coluna_lancamento + str(linha_especifica)] = protocolo
    coluna_status = 'C'  
    if linha > ws.max_row:
        ws[coluna_status + str(linha_especifica)] = status
    else:
        ws[coluna_status + str(linha_especifica)] = status
    wb.save(caminho_do_arquivo)
    wb.close()

    