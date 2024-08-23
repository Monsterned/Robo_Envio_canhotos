import pandas as pd
from openpyxl import load_workbook
import os
import numpy as np
import pyautogui
import ctypes
from datetime import datetime
import pyautogui
from selenium.webdriver.common.action_chains import ActionChains
import shutil
from datetime import datetime, timedelta
import pyperclip

caminho = os.getcwd() 
caminho_sistema = caminho.replace("C", "T", 1)

def click_image(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def confirmacao_na_tela(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Imagem foi confirmada na tela.")
                break
        except Exception as e:
            print("Imagem não confirmada na tela. Aguardando...")
        pyautogui.sleep(1)

def salvar_lancamento(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    protocolo = None  # Inicializa a variável protocolo
    
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Imagem foi confirmada na tela.")
        except Exception as e:
            print("Imagem não confirmada na tela. Aguardando...")
            click_info_lancamento('codigo_lancamento.png')
            pyautogui.click(button='right')
            pyautogui.sleep(1)
            click_image('copy.png')
            pyautogui.sleep(0.5)
            try:
                text = pyperclip.paste()
                protocolo = int(text)
                print("Número do LANCAMENTO:", protocolo)
            except ValueError:
                print("O conteúdo copiado não é um número válido.")
            except Exception as e:
                print("Ocorreu um erro:", str(e))
            break
        pyautogui.sleep(1)
    
    return protocolo  # Retorna o valor de protocolo

def click_info_lancamento(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.moveTo(center_x, center_y)  # Movendo o cursor para a posição da imagem               
                pyautogui.moveRel(60, 0)  # Movendo o cursor para cima
                pyautogui.click()  # Clicando no local da imagem
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)


def click_info_cte(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.moveTo(center_x, center_y)  # Movendo o cursor para a posição da imagem               
                pyautogui.moveRel(40, 0)  # Movendo o cursor para cima
                pyautogui.click()  # Clicando no local da imagem
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

Planilha_canhoto = pd.read_excel("Base_canhotos.xlsx")
#print(Planilha_canhoto)

status = datetime.now()
status = status.strftime("%d/%m")
status = f'Enviado {status}'
#print(status)

click_image('botao_util.png')
click_image('botao_util_acessar.png')
click_image('botao_protocolo.png')
click_image('botao_protocolo_lancamento.png')
click_image('incluir.png')
click_image('janela_geral.png')
confirmacao_na_tela('codigo_automatico.png')
click_info_lancamento('usuario_destino.png')
pyautogui.write('FOGACA2')
pyautogui.press('tab')
click_info_lancamento('responsavel.png')
pyautogui.write('JHENIFER FRANCIELE')
pyautogui.press('tab')
click_info_lancamento('tipo.png')
pyautogui.write('3')
pyautogui.press('tab')
click_info_lancamento('cliente_fornecedor.png')
pyautogui.write('630311')
pyautogui.press('tab')
click_info_lancamento('mensagem_envio.png')
pyautogui.write('PROTOCOLO DE ENTREGA DE CANHOTOS CLIENTE CACAU SHOW (OBS. INFORMAR O N° DO PROTOCOLO NA FATURA)')
pyautogui.sleep(5)
pyautogui.press('tab')
click_info_lancamento('municipio.png')
pyautogui.write('7932')
pyautogui.press('tab')
click_image('salvar.png')
protocolo = salvar_lancamento('codigo_automatico.png')
click_image('janela_documento.png')

numero_linhas = len(Planilha_canhoto)
#print(numero_linhas)
linha_especifica = 1
for i, linha in enumerate(Planilha_canhoto.index):
    cte = Planilha_canhoto.loc[linha, "Ct-e Online"]
    cte = int(cte)
    #print(cte)
    linha_especifica += 1 
    caminho_do_arquivo = 'Base_canhotos.xlsx'
    nome_da_aba = 'Plan1'
    wb = load_workbook(caminho_do_arquivo)
    ws = wb[nome_da_aba]
    coluna_lancamento = 'B'  
    if linha > ws.max_row:
        ws[coluna_lancamento + str(linha_especifica)] = protocolo
    else:
        ws[coluna_lancamento + str(linha_especifica)] = protocolo
    coluna_status = 'C'  
    if linha > ws.max_row:
        ws[coluna_status + str(linha_especifica)] = status
    else:
        ws[coluna_status + str(linha_especifica)] = status
    wb.save(caminho_do_arquivo)
    wb.close()


    if i == 0:
        click_image('pasta_amarela.png')
    else:
        click_image('pasta_amarela_marcada.png')
    pyautogui.sleep(2)
    pyautogui.press('tab')
    click_info_cte('cte_filial.png')
    pyautogui.write('1')
    pyautogui.press('tab')
    click_info_cte('cte_serie.png')
    pyautogui.write('2')
    pyautogui.press('tab')
    click_info_cte('cte_documento.png')
    pyautogui.write(str(cte))
    pyautogui.press('tab')
    click_image('setinha_verde.png')
    pyautogui.sleep(1)
    confirmacao_na_tela('confirmacao_inclusao.png')

click_image('janela_geral.png')
click_image('enviar.png')
click_image('ok.png')
pyautogui.sleep(5)
click_image('sair.png')
click_image('yes.png')


